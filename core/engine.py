
import os
import re
import mimetypes
import smtplib
from io import BytesIO
from datetime import datetime
from email.message import EmailMessage

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# EXCEPCIONES
# ============================================================

class DataControlHubSchemaError(Exception):
    """Error controlado cuando el Excel no cumple el esquema mínimo."""

    def __init__(self, message, schema_issues=None):
        super().__init__(message)
        self.schema_issues = schema_issues if schema_issues is not None else pd.DataFrame()


# ============================================================
# FUNCIONES DE LIMPIEZA Y NORMALIZACIÓN
# ============================================================

def clean_text(value):
    if pd.isna(value):
        return ""
    value = str(value).strip()
    value = re.sub(r"\s+", " ", value)
    return value


def normalize_colname(col):
    col = clean_text(col).lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ñ": "n", "ü": "u"
    }
    for old, new in replacements.items():
        col = col.replace(old, new)
    col = re.sub(r"[^a-z0-9]+", "_", col).strip("_")
    return col


def normalize_id(value):
    value = clean_text(value).upper()
    if value.endswith(".0"):
        value = value[:-2]
    return value


def normalize_status(value):
    value = clean_text(value).lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ñ": "n", "ü": "u"
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"\s+", " ", value).strip()

    status_map = {
        "activo": "Activo",
        "activa": "Activo",
        "vigente": "Activo",
        "desvinculado": "Desvinculado",
        "desvinculada": "Desvinculado",
        "finiquitado": "Desvinculado",
        "finiquitada": "Desvinculado",
        "retirado": "Desvinculado",
        "retirada": "Desvinculado",
        "suspendido": "Suspendido",
        "suspendida": "Suspendido",
        "trasladado": "Trasladado",
        "trasladada": "Trasladado",
        "inactivo": "Inactivo",
        "inactiva": "Inactivo",
        "asignado": "Asignado",
        "asignada": "Asignado",
        "disponible": "Disponible",
        "pendiente devolucion": "Pendiente devolución",
        "pendiente devolución": "Pendiente devolución",
        "recuperado": "Recuperado",
        "recuperada": "Recuperado",
        "en reparacion": "En reparación",
        "en reparación": "En reparación",
        "danado": "Dañado",
        "dañado": "Dañado",
        "danada": "Dañado",
        "dañada": "Dañado",
        "dado de baja": "Dado de baja",
        "dada de baja": "Dado de baja",
        "baja": "Dado de baja",
        "extraviado": "Extraviado",
        "extraviada": "Extraviado",
        "robado": "Robado",
        "robada": "Robado",
        "operativo": "Operativo",
        "operativa": "Operativo",
        "requiere revision": "Requiere revisión",
        "requiere revisión": "Requiere revisión",
        "requiere reparacion": "Requiere reparación",
        "requiere reparación": "Requiere reparación",
        "en observacion": "En observación",
        "en observación": "En observación",
        "no reparable": "No reparable",
        "baja sugerida": "Baja sugerida",
        "bueno": "Bueno",
        "buena": "Bueno",
        "regular": "Regular",
        "malo": "Malo",
        "mala": "Malo",
    }
    return status_map.get(value, clean_text(value).title())


def to_numeric_safe(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def detect_sheet(xls, candidates):
    sheet_map = {normalize_colname(s): s for s in xls.sheet_names}
    for candidate in candidates:
        key = normalize_colname(candidate)
        if key in sheet_map:
            return sheet_map[key]
    return None


def standardize_columns(df, aliases):
    df = df.copy()
    original_cols = list(df.columns)
    normalized_to_original = {normalize_colname(c): c for c in original_cols}

    rename_map = {}
    for standard_col, possible_names in aliases.items():
        for name in possible_names:
            norm = normalize_colname(name)
            if norm in normalized_to_original:
                rename_map[normalized_to_original[norm]] = standard_col
                break

    df = df.rename(columns=rename_map)
    df.columns = [normalize_colname(c) if c not in aliases.keys() else c for c in df.columns]
    return df


def first_existing_col(df, cols):
    for col in cols:
        if col in df.columns:
            return col
    return None


def rule_active(reglas_df, rule_code, default=True):
    if reglas_df is None or reglas_df.empty or "codigo_regla" not in reglas_df.columns:
        return default
    row = reglas_df[reglas_df["codigo_regla"] == rule_code]
    if row.empty:
        return default
    if "activa" not in row.columns:
        return default
    val = clean_text(row.iloc[0]["activa"]).lower()
    return val in ["si", "sí", "s", "yes", "y", "true", "1"]


# ============================================================
# ALIASES
# ============================================================

ALIASES_PERSONAS = {
    "id_empleado": ["id_empleado", "id empleado", "id_colaborador", "id trabajador", "legajo", "employee_id"],
    "rut": ["rut", "run", "documento"],
    "nombre": ["nombre", "nombre completo", "full_name", "colaborador"],
    "correo": ["correo", "email", "mail"],
    "area": ["area", "área", "departamento", "unidad"],
    "cargo": ["cargo", "puesto", "position"],
    "estado_persona": ["estado_persona", "estado persona", "estado empleado", "estado contrato", "estado_colaborador"],
    "fecha_estado": ["fecha_estado", "fecha estado", "fecha cambio estado", "status_date"],
}

ALIASES_ACTIVOS = {
    "codigo_activo": ["codigo_activo", "código activo", "codigo activo", "asset_code", "id_activo"],
    "numero_serie": ["numero_serie", "número serie", "numero serie", "serial_number", "serie"],
    "tipo_activo": ["tipo_activo", "tipo activo", "asset_type", "tipo dispositivo"],
    "marca": ["marca", "brand"],
    "modelo": ["modelo", "model"],
    "id_empleado_asignado": ["id_empleado_asignado", "id empleado asignado", "id_asignado", "employee_id", "id_empleado"],
    "estado_activo": ["estado_activo", "estado activo", "asset_status", "estado dispositivo"],
    "fecha_asignacion": ["fecha_asignacion", "fecha asignacion", "fecha asignación", "assigned_date"],
    "ubicacion": ["ubicacion", "ubicación", "location"],
}

ALIASES_ARTICULOS = {
    "codigo_activo": ["codigo_activo", "código activo", "codigo activo", "asset_code", "id_activo"],
    "tipo_activo": ["tipo_activo", "tipo activo", "asset_type"],
    "fecha_compra": ["fecha_compra", "fecha compra", "purchase_date"],
    "estado_actual": ["estado_actual", "estado actual", "estado_articulo", "estado articulo"],
    "condicion_fisica": ["condicion_fisica", "condición física", "condicion fisica", "physical_condition"],
    "cantidad_fallas": ["cantidad_fallas", "cantidad fallas", "fallas", "failure_count"],
    "reparaciones_previas": ["reparaciones_previas", "reparaciones previas", "repair_count"],
    "costo_reparacion": ["costo_reparacion", "costo reparación", "costo reparacion", "estimated_repair_cost"],
    "costo_reposicion": ["costo_reposicion", "costo reposición", "costo reposicion", "replacement_cost"],
    "responsable": ["responsable", "owner", "area_responsable"],
}

ALIASES_REGLAS = {
    "codigo_regla": ["codigo_regla", "código regla", "codigo", "regla", "rule_code"],
    "orden_prioridad": ["orden_prioridad", "orden prioridad", "prioridad", "priority_order"],
    "nombre_regla": ["nombre_regla", "nombre regla", "tipo_alerta", "rule_name"],
    "criticidad": ["criticidad", "severity"],
    "prioridad_operacional": ["prioridad_operacional", "prioridad operacional"],
    "sla_dias": ["sla_dias", "sla dias", "sla"],
    "accion_sugerida": ["accion_sugerida", "acción sugerida", "accion sugerida"],
    "responsable_sugerido": ["responsable_sugerido", "responsable sugerido"],
    "activa": ["activa", "activo", "is_active"],
}


# ============================================================
# REGLAS OFICIALES
# ============================================================

def default_rules():
    return pd.DataFrame([
        {"codigo_regla":"R005","orden_prioridad":1,"nombre_regla":"Activo asignado a persona inexistente","criticidad":"Crítica","prioridad_operacional":"Urgente","sla_dias":3,"accion_sugerida":"Revisar manualmente asignación","responsable_sugerido":"Activos / RRHH","activa":"Sí"},
        {"codigo_regla":"R001","orden_prioridad":2,"nombre_regla":"Persona desvinculada con activo asignado","criticidad":"Crítica","prioridad_operacional":"Urgente","sla_dias":3,"accion_sugerida":"Recuperar activo","responsable_sugerido":"Activos / TI / RECO","activa":"Sí"},
        {"codigo_regla":"R002","orden_prioridad":3,"nombre_regla":"Persona suspendida con activo asignado","criticidad":"Alta","prioridad_operacional":"Alta","sla_dias":7,"accion_sugerida":"Revisar continuidad operacional","responsable_sugerido":"RRHH / Jefatura / TI","activa":"Sí"},
        {"codigo_regla":"R003","orden_prioridad":4,"nombre_regla":"Persona inactiva con activo asignado","criticidad":"Alta","prioridad_operacional":"Alta","sla_dias":7,"accion_sugerida":"Recuperar o revisar activo","responsable_sugerido":"Activos / Operaciones","activa":"Sí"},
        {"codigo_regla":"R006","orden_prioridad":5,"nombre_regla":"Activo dañado pero sigue asignado","criticidad":"Alta","prioridad_operacional":"Alta","sla_dias":7,"accion_sugerida":"Retirar o reparar activo","responsable_sugerido":"TI / Bodega / Proveedor","activa":"Sí"},
        {"codigo_regla":"R004","orden_prioridad":6,"nombre_regla":"Persona trasladada con activo asignado","criticidad":"Media","prioridad_operacional":"Media","sla_dias":15,"accion_sugerida":"Revisar reasignación","responsable_sugerido":"Jefatura / Activos","activa":"Sí"},
        {"codigo_regla":"R007","orden_prioridad":7,"nombre_regla":"Activo con más de 2 reparaciones previas","criticidad":"Alta","prioridad_operacional":"Alta","sla_dias":7,"accion_sugerida":"Evaluar baja","responsable_sugerido":"TI / Finanzas / Activos","activa":"Sí"},
        {"codigo_regla":"R008","orden_prioridad":8,"nombre_regla":"Costo de reparación supera umbral","criticidad":"Alta","prioridad_operacional":"Alta","sla_dias":7,"accion_sugerida":"Evaluar baja","responsable_sugerido":"TI / Compras / Finanzas","activa":"Sí"},
        {"codigo_regla":"R009","orden_prioridad":9,"nombre_regla":"Activo con condición física mala","criticidad":"Alta","prioridad_operacional":"Alta","sla_dias":7,"accion_sugerida":"Revisar reparación o baja","responsable_sugerido":"Activos / Bodega / TI","activa":"Sí"},
        {"codigo_regla":"R010","orden_prioridad":10,"nombre_regla":"Activo en reparación","criticidad":"Media","prioridad_operacional":"Media","sla_dias":15,"accion_sugerida":"Hacer seguimiento","responsable_sugerido":"TI / Proveedor","activa":"Sí"},
    ])


# ============================================================
# LECTURA Y PREPARACIÓN
# ============================================================

def load_source_excel(source_file):
    xls = pd.ExcelFile(source_file)

    sheet_personas = detect_sheet(xls, ["personas_rrhh", "personas_procesadas", "personas", "rrhh"])
    sheet_activos = detect_sheet(xls, ["activos_dispositivos", "activos_procesados", "activos", "dispositivos"])
    sheet_articulos = detect_sheet(xls, ["articulos_reparacion_baja", "articulos_procesados", "articulos", "reparacion_baja"])
    sheet_reglas = detect_sheet(xls, ["reglas_mvp", "reglas_aplicadas", "business_rules"])

    missing_sheets = []
    if sheet_personas is None:
        missing_sheets.append("personas_rrhh")
    if sheet_activos is None:
        missing_sheets.append("activos_dispositivos")
    if sheet_articulos is None:
        missing_sheets.append("articulos_reparacion_baja")

    if missing_sheets:
        issues = pd.DataFrame([{
            "hoja": sheet,
            "columna_requerida": "-",
            "problema": "Hoja obligatoria no encontrada"
        } for sheet in missing_sheets])
        raise DataControlHubSchemaError("Faltan hojas mínimas en el Excel fuente.", issues)

    personas_raw = pd.read_excel(source_file, sheet_name=sheet_personas)
    activos_raw = pd.read_excel(source_file, sheet_name=sheet_activos)
    articulos_raw = pd.read_excel(source_file, sheet_name=sheet_articulos)
    reglas_raw = pd.read_excel(source_file, sheet_name=sheet_reglas) if sheet_reglas else pd.DataFrame()

    return personas_raw, activos_raw, articulos_raw, reglas_raw


def standardize_and_validate_schema(personas_raw, activos_raw, articulos_raw, reglas_raw):
    personas = standardize_columns(personas_raw, ALIASES_PERSONAS)
    activos = standardize_columns(activos_raw, ALIASES_ACTIVOS)
    articulos = standardize_columns(articulos_raw, ALIASES_ARTICULOS)
    reglas = standardize_columns(reglas_raw, ALIASES_REGLAS) if reglas_raw is not None and not reglas_raw.empty else pd.DataFrame()

    required_personas = ["id_empleado", "nombre", "estado_persona"]
    required_activos = ["codigo_activo", "tipo_activo", "id_empleado_asignado", "estado_activo"]
    required_articulos = ["codigo_activo", "estado_actual", "condicion_fisica", "reparaciones_previas", "costo_reparacion", "costo_reposicion"]

    schema_issues = []
    for col in required_personas:
        if col not in personas.columns:
            schema_issues.append({"hoja": "personas_rrhh", "columna_requerida": col, "problema": "Columna obligatoria no encontrada"})
    for col in required_activos:
        if col not in activos.columns:
            schema_issues.append({"hoja": "activos_dispositivos", "columna_requerida": col, "problema": "Columna obligatoria no encontrada"})
    for col in required_articulos:
        if col not in articulos.columns:
            schema_issues.append({"hoja": "articulos_reparacion_baja", "columna_requerida": col, "problema": "Columna obligatoria no encontrada"})

    if schema_issues:
        raise DataControlHubSchemaError(
            "El Excel no cumple el esquema mínimo.",
            pd.DataFrame(schema_issues)
        )

    return personas, activos, articulos, reglas


def normalize_data(personas, activos, articulos, reglas):
    personas = personas.copy()
    activos = activos.copy()
    articulos = articulos.copy()
    reglas = reglas.copy()

    personas["id_empleado"] = personas["id_empleado"].map(normalize_id)
    personas["estado_persona"] = personas["estado_persona"].map(normalize_status)
    personas["nombre"] = personas["nombre"].map(clean_text)

    for col in ["correo", "area", "cargo", "rut"]:
        if col in personas.columns:
            if col == "correo":
                personas[col] = personas[col].map(lambda x: clean_text(x).lower())
            else:
                personas[col] = personas[col].map(clean_text)

    activos["codigo_activo"] = activos["codigo_activo"].map(normalize_id)
    activos["id_empleado_asignado"] = activos["id_empleado_asignado"].map(normalize_id)
    activos["estado_activo"] = activos["estado_activo"].map(normalize_status)
    activos["tipo_activo"] = activos["tipo_activo"].map(clean_text)

    for col in ["numero_serie", "marca", "modelo", "ubicacion"]:
        if col in activos.columns:
            activos[col] = activos[col].map(clean_text)

    articulos["codigo_activo"] = articulos["codigo_activo"].map(normalize_id)
    articulos["estado_actual"] = articulos["estado_actual"].map(normalize_status)
    articulos["condicion_fisica"] = articulos["condicion_fisica"].map(normalize_status)
    articulos["reparaciones_previas"] = to_numeric_safe(articulos["reparaciones_previas"])
    articulos["costo_reparacion"] = to_numeric_safe(articulos["costo_reparacion"])
    articulos["costo_reposicion"] = to_numeric_safe(articulos["costo_reposicion"])

    if not reglas.empty and "codigo_regla" in reglas.columns:
        reglas["codigo_regla"] = reglas["codigo_regla"].map(lambda x: clean_text(x).upper())
        if "activa" in reglas.columns:
            reglas["activa"] = reglas["activa"].map(clean_text)

    return personas, activos, articulos, reglas


# ============================================================
# VALIDACIONES
# ============================================================

def validate_data_quality(personas, activos, articulos):
    validaciones = []

    def add_validation(tipo, hoja, registro, problema, severidad="Media"):
        validaciones.append({
            "tipo_validacion": tipo,
            "hoja": hoja,
            "registro": registro,
            "problema": problema,
            "severidad": severidad
        })

    for idx, row in personas.iterrows():
        if not row.get("id_empleado"):
            add_validation("Campo vacío", "personas_rrhh", idx + 2, "id_empleado vacío", "Crítica")
        if not row.get("estado_persona"):
            add_validation("Campo vacío", "personas_rrhh", row.get("id_empleado", idx + 2), "estado_persona vacío", "Alta")

    duplicated_personas = personas[personas["id_empleado"].duplicated(keep=False) & (personas["id_empleado"] != "")]
    for _, row in duplicated_personas.iterrows():
        add_validation("Duplicado", "personas_rrhh", row["id_empleado"], "id_empleado duplicado", "Alta")

    for idx, row in activos.iterrows():
        if not row.get("codigo_activo"):
            add_validation("Campo vacío", "activos_dispositivos", idx + 2, "codigo_activo vacío", "Crítica")
        if row.get("estado_activo") == "Asignado" and not row.get("id_empleado_asignado"):
            add_validation("Campo vacío", "activos_dispositivos", row.get("codigo_activo", idx + 2), "Activo asignado sin id_empleado_asignado", "Crítica")

    duplicated_activos = activos[activos["codigo_activo"].duplicated(keep=False) & (activos["codigo_activo"] != "")]
    for _, row in duplicated_activos.iterrows():
        add_validation("Duplicado", "activos_dispositivos", row["codigo_activo"], "codigo_activo duplicado", "Alta")

    person_ids = set(personas["id_empleado"].dropna())
    activos_asignados = activos[activos["estado_activo"].isin(["Asignado", "Pendiente devolución"])]

    for _, row in activos_asignados.iterrows():
        id_asig = row.get("id_empleado_asignado")
        if id_asig and id_asig not in person_ids:
            add_validation("Relación inexistente", "activos_dispositivos", row["codigo_activo"], f"id_empleado_asignado {id_asig} no existe en RRHH", "Crítica")

    asset_codes = set(activos["codigo_activo"].dropna())
    for _, row in articulos.iterrows():
        if row.get("codigo_activo") and row.get("codigo_activo") not in asset_codes:
            add_validation("Relación inexistente", "articulos_reparacion_baja", row["codigo_activo"], "codigo_activo no existe en activos_dispositivos", "Alta")
        if row.get("costo_reposicion", 0) <= 0 and row.get("costo_reparacion", 0) > 0:
            add_validation("Costo inválido", "articulos_reparacion_baja", row.get("codigo_activo"), "costo_reposicion vacío o cero con costo_reparacion mayor a cero", "Alta")

    validacion_datos = pd.DataFrame(validaciones)
    if validacion_datos.empty:
        validacion_datos = pd.DataFrame([{
            "tipo_validacion": "Sin observaciones",
            "hoja": "-",
            "registro": "-",
            "problema": "No se detectaron problemas de calidad de datos",
            "severidad": "Informativa"
        }])

    return validacion_datos


# ============================================================
# MOTOR DE REGLAS
# ============================================================

def prepare_rules(reglas):
    reglas_aplicadas = default_rules()

    if reglas is not None and not reglas.empty and "codigo_regla" in reglas.columns and "activa" in reglas.columns:
        for _, row in reglas.iterrows():
            code = clean_text(row.get("codigo_regla")).upper()
            if code in set(reglas_aplicadas["codigo_regla"]):
                reglas_aplicadas.loc[reglas_aplicadas["codigo_regla"] == code, "activa"] = clean_text(row.get("activa", "Sí"))

    return reglas_aplicadas


def cross_sources(personas, activos, articulos):
    personas_lookup = personas.drop_duplicates("id_empleado", keep="last")
    articulos_lookup = articulos.drop_duplicates("codigo_activo", keep="last")

    persona_cols = ["id_empleado", "nombre", "estado_persona"] + [
        c for c in ["rut", "correo", "area", "cargo"] if c in personas_lookup.columns
    ]

    articulo_cols = ["codigo_activo", "estado_actual", "condicion_fisica", "reparaciones_previas", "costo_reparacion", "costo_reposicion"] + [
        c for c in ["responsable"] if c in articulos_lookup.columns
    ]

    base = activos.merge(
        personas_lookup[persona_cols],
        how="left",
        left_on="id_empleado_asignado",
        right_on="id_empleado"
    )

    base = base.merge(
        articulos_lookup[articulo_cols],
        how="left",
        on="codigo_activo"
    )

    return base


def apply_rules(base, reglas_aplicadas, umbral_costo_reparacion=0.70):
    rules_meta = reglas_aplicadas.set_index("codigo_regla").to_dict("index")

    def is_active(rule_code):
        return rule_active(reglas_aplicadas, rule_code, default=True)

    def get_rule_meta(rule_code):
        meta = rules_meta.get(rule_code, {})
        return {
            "orden_prioridad": meta.get("orden_prioridad", np.nan),
            "tipo_alerta": meta.get("nombre_regla", rule_code),
            "criticidad": meta.get("criticidad", "Media"),
            "prioridad_operacional": meta.get("prioridad_operacional", "Media"),
            "sla_dias": meta.get("sla_dias", 15),
            "accion_sugerida": meta.get("accion_sugerida", "Revisar"),
            "responsable_sugerido": meta.get("responsable_sugerido", "Pendiente asignar"),
        }

    def evaluate_row(row):
        estado_persona = row.get("estado_persona", "")
        estado_activo = row.get("estado_activo", "")
        estado_actual = row.get("estado_actual", "")
        condicion = row.get("condicion_fisica", "")
        id_empleado = row.get("id_empleado_asignado", "")
        persona_encontrada = pd.notna(row.get("id_empleado")) and clean_text(row.get("id_empleado")) != ""

        reparaciones = row.get("reparaciones_previas", 0) or 0
        costo_rep = row.get("costo_reparacion", 0) or 0
        costo_repo = row.get("costo_reposicion", 0) or 0

        activo_asignado = estado_activo in ["Asignado", "Pendiente devolución"]

        if is_active("R005") and activo_asignado and id_empleado and not persona_encontrada:
            return "R005", f"Activo asignado a ID {id_empleado}, pero ese ID no existe en RRHH."
        if is_active("R001") and activo_asignado and estado_persona == "Desvinculado":
            return "R001", "Persona desvinculada mantiene activo asignado o pendiente de devolución."
        if is_active("R002") and activo_asignado and estado_persona == "Suspendido":
            return "R002", "Persona suspendida mantiene activo asignado o pendiente de devolución."
        if is_active("R003") and activo_asignado and estado_persona == "Inactivo":
            return "R003", "Persona inactiva mantiene activo asignado o pendiente de devolución."
        if is_active("R006") and activo_asignado and estado_actual == "Dañado":
            return "R006", "Activo dañado sigue asignado."
        if is_active("R004") and activo_asignado and estado_persona == "Trasladado":
            return "R004", "Persona trasladada mantiene activo asignado; revisar si corresponde reasignación."
        if is_active("R007") and reparaciones > 2:
            return "R007", f"Activo registra {int(reparaciones)} reparaciones previas."
        if is_active("R008") and costo_repo > 0 and costo_rep > costo_repo * umbral_costo_reparacion:
            return "R008", f"Costo de reparación ({costo_rep:,.0f}) supera el {umbral_costo_reparacion:.0%} del costo de reposición ({costo_repo:,.0f})."
        if is_active("R009") and condicion == "Malo":
            return "R009", "Activo con condición física mala."
        if is_active("R010") and estado_actual == "En reparación":
            return "R010", "Activo se encuentra en reparación y requiere seguimiento."

        return "", ""

    base = base.copy()
    evaluated = base.apply(evaluate_row, axis=1, result_type="expand")
    base["codigo_regla"] = evaluated[0]
    base["motivo_alerta"] = evaluated[1]
    base["debe_atender"] = np.where(base["codigo_regla"] != "", "Sí", "No")

    alertas = base[base["debe_atender"] == "Sí"].copy()

    for col in ["orden_prioridad", "tipo_alerta", "criticidad", "prioridad_operacional", "sla_dias", "accion_sugerida", "responsable_sugerido"]:
        alertas[col] = alertas["codigo_regla"].map(lambda r: get_rule_meta(r)[col])

    alertas["id_alerta"] = (
        alertas["codigo_regla"].astype(str)
        + "-"
        + alertas["codigo_activo"].astype(str)
        + "-"
        + alertas["id_empleado_asignado"].astype(str).replace("", "SIN_ID")
    )

    alertas["fecha_generacion"] = pd.Timestamp.today().normalize()
    alertas["vigente_en_ultima_carga"] = "Sí"

    cols_alertas = [
        "id_alerta", "codigo_regla", "orden_prioridad", "tipo_alerta", "criticidad",
        "prioridad_operacional", "sla_dias", "id_empleado_asignado", "nombre",
        "estado_persona", "codigo_activo", "tipo_activo", "estado_activo",
        "estado_actual", "condicion_fisica", "reparaciones_previas",
        "costo_reparacion", "costo_reposicion", "motivo_alerta",
        "accion_sugerida", "responsable_sugerido", "fecha_generacion",
        "vigente_en_ultima_carga"
    ]
    cols_alertas = [c for c in cols_alertas if c in alertas.columns]

    alertas_generadas = alertas[cols_alertas].sort_values(["orden_prioridad", "id_alerta"]).reset_index(drop=True)
    return alertas_generadas


# ============================================================
# PERSISTENCIA
# ============================================================

def read_previous_file(previous_file):
    if previous_file is None:
        return pd.DataFrame(), pd.DataFrame()

    try:
        prev_xls = pd.ExcelFile(previous_file)
        sheet_acciones = detect_sheet(prev_xls, ["acciones_pendientes", "gestion_alertas"])
        sheet_historial = detect_sheet(prev_xls, ["historial_alertas"])

        gestion_anterior = pd.read_excel(previous_file, sheet_name=sheet_acciones) if sheet_acciones else pd.DataFrame()
        historial_anterior = pd.read_excel(previous_file, sheet_name=sheet_historial) if sheet_historial else pd.DataFrame()

        if not gestion_anterior.empty:
            gestion_anterior.columns = [normalize_colname(c) for c in gestion_anterior.columns]
        if not historial_anterior.empty:
            historial_anterior.columns = [normalize_colname(c) for c in historial_anterior.columns]

        return gestion_anterior, historial_anterior

    except Exception:
        return pd.DataFrame(), pd.DataFrame()


def generate_actions(alertas_generadas, gestion_anterior):
    acciones = alertas_generadas.copy()

    gestion_cols = [
        "estado_gestion",
        "responsable_manual",
        "fecha_compromiso",
        "fecha_cierre",
        "decision_final",
        "comentario_manual",
        "evidencia",
        "usuario_gestion"
    ]

    acciones["estado_gestion"] = "Nueva"
    acciones["responsable_manual"] = ""
    acciones["fecha_compromiso"] = ""
    acciones["fecha_cierre"] = ""
    acciones["decision_final"] = ""
    acciones["comentario_manual"] = ""
    acciones["evidencia"] = ""
    acciones["usuario_gestion"] = ""

    if not gestion_anterior.empty and "id_alerta" in gestion_anterior.columns:
        keep_cols = ["id_alerta"] + [c for c in gestion_cols if c in gestion_anterior.columns]
        prev_keep = gestion_anterior[keep_cols].drop_duplicates("id_alerta", keep="last")

        acciones = acciones.merge(prev_keep, how="left", on="id_alerta", suffixes=("", "_anterior"))

        for col in gestion_cols:
            prev_col = f"{col}_anterior"
            if prev_col in acciones.columns:
                acciones[col] = acciones[prev_col].where(
                    acciones[prev_col].notna() & (acciones[prev_col].astype(str) != ""),
                    acciones[col]
                )
                acciones = acciones.drop(columns=[prev_col])

        ids_previos = set(gestion_anterior["id_alerta"].astype(str))
        acciones["situacion_alerta"] = np.where(
            acciones["id_alerta"].astype(str).isin(ids_previos),
            "Persistente",
            "Nueva detección"
        )
    else:
        acciones["situacion_alerta"] = "Nueva detección"

    fecha_ejecucion = pd.Timestamp.today().normalize()
    acciones["fecha_ultima_deteccion"] = fecha_ejecucion
    acciones["fecha_generacion"] = pd.to_datetime(acciones["fecha_generacion"], errors="coerce")
    acciones["dias_abierta"] = (fecha_ejecucion - acciones["fecha_generacion"]).dt.days
    acciones["dias_abierta"] = acciones["dias_abierta"].fillna(0).astype(int)
    acciones["sla_dias"] = pd.to_numeric(acciones["sla_dias"], errors="coerce").fillna(15).astype(int)
    acciones["alerta_vencida"] = np.where(
        (~acciones["estado_gestion"].isin(["Resuelta", "Descartada"])) & (acciones["dias_abierta"] > acciones["sla_dias"]),
        "Sí",
        "No"
    )

    cols_acciones = [
        "id_alerta", "situacion_alerta", "codigo_regla", "orden_prioridad", "criticidad",
        "prioridad_operacional", "sla_dias", "alerta_vencida", "dias_abierta",
        "tipo_alerta", "id_empleado_asignado", "nombre", "estado_persona",
        "codigo_activo", "tipo_activo", "estado_activo", "estado_actual",
        "motivo_alerta", "accion_sugerida", "responsable_sugerido",
        "estado_gestion", "responsable_manual", "fecha_compromiso", "fecha_cierre",
        "decision_final", "comentario_manual", "evidencia", "usuario_gestion",
        "fecha_generacion", "fecha_ultima_deteccion", "vigente_en_ultima_carga"
    ]
    cols_acciones = [c for c in cols_acciones if c in acciones.columns]

    acciones_pendientes = acciones[cols_acciones].sort_values(
        ["orden_prioridad", "alerta_vencida", "id_alerta"],
        ascending=[True, False, True]
    ).reset_index(drop=True)

    return acciones_pendientes


def generate_history(acciones_pendientes, historial_anterior):
    fecha_ejecucion = pd.Timestamp.today().normalize()

    base_cols = [
        "id_alerta", "codigo_regla", "tipo_alerta", "criticidad", "orden_prioridad",
        "codigo_activo", "id_empleado_asignado", "estado_gestion",
        "fecha_generacion", "fecha_ultima_deteccion", "vigente_en_ultima_carga"
    ]
    base_cols = [c for c in base_cols if c in acciones_pendientes.columns]

    historial_actual = acciones_pendientes[base_cols].copy()
    historial_actual["primera_fecha_detectada"] = historial_actual.get("fecha_generacion", fecha_ejecucion)
    historial_actual["ultima_fecha_detectada"] = fecha_ejecucion
    historial_actual["veces_detectada"] = 1

    if historial_anterior is not None and not historial_anterior.empty and "id_alerta" in historial_anterior.columns:
        for col in ["primera_fecha_detectada", "ultima_fecha_detectada", "veces_detectada"]:
            if col not in historial_anterior.columns:
                historial_anterior[col] = ""

        prev = historial_anterior.drop_duplicates("id_alerta", keep="last").copy()
        current_ids = set(historial_actual["id_alerta"].astype(str))

        hist = historial_actual.merge(
            prev[["id_alerta", "primera_fecha_detectada", "veces_detectada"]],
            how="left",
            on="id_alerta",
            suffixes=("", "_anterior")
        )

        hist["primera_fecha_detectada"] = hist["primera_fecha_detectada_anterior"].where(
            hist["primera_fecha_detectada_anterior"].notna() & (hist["primera_fecha_detectada_anterior"].astype(str) != ""),
            hist["primera_fecha_detectada"]
        )

        hist["veces_detectada_anterior"] = pd.to_numeric(hist["veces_detectada_anterior"], errors="coerce").fillna(0).astype(int)
        hist["veces_detectada"] = hist["veces_detectada_anterior"] + 1
        hist["ultima_fecha_detectada"] = fecha_ejecucion
        hist["vigente_en_ultima_carga"] = "Sí"

        drop_cols = [c for c in hist.columns if c.endswith("_anterior")]
        hist = hist.drop(columns=drop_cols)

        prev_no_vigentes = prev[~prev["id_alerta"].astype(str).isin(current_ids)].copy()
        if not prev_no_vigentes.empty:
            prev_no_vigentes["vigente_en_ultima_carga"] = "No"
            prev_no_vigentes["estado_actual"] = "No vigente"

            for col in hist.columns:
                if col not in prev_no_vigentes.columns:
                    prev_no_vigentes[col] = ""
            prev_no_vigentes = prev_no_vigentes[hist.columns]

            historial_alertas = pd.concat([hist, prev_no_vigentes], ignore_index=True)
        else:
            historial_alertas = hist.copy()
    else:
        historial_alertas = historial_actual.copy()

    if "estado_actual" not in historial_alertas.columns:
        historial_alertas["estado_actual"] = historial_alertas.get("estado_gestion", "")

    historial_alertas["estado_actual"] = np.where(
        historial_alertas["vigente_en_ultima_carga"] == "No",
        "No vigente",
        historial_alertas.get("estado_gestion", historial_alertas["estado_actual"])
    )

    historial_cols = [
        "id_alerta", "codigo_regla", "tipo_alerta", "criticidad", "orden_prioridad",
        "codigo_activo", "id_empleado_asignado", "primera_fecha_detectada",
        "ultima_fecha_detectada", "veces_detectada", "estado_actual",
        "vigente_en_ultima_carga"
    ]
    historial_cols = [c for c in historial_cols if c in historial_alertas.columns]

    return historial_alertas[historial_cols].sort_values(
        ["vigente_en_ultima_carga", "orden_prioridad", "id_alerta"],
        ascending=[False, True, True]
    ).reset_index(drop=True)


# ============================================================
# DASHBOARD Y EXPORTACIÓN
# ============================================================

def count_value(df, col, value):
    if df is None or df.empty or col not in df.columns:
        return 0
    return int((df[col] == value).sum())


def generate_summaries(personas, activos, articulos, validacion_datos, alertas_generadas, acciones_pendientes, historial_alertas, archivo_fuente_nombre, umbral):
    dashboard_data = [
        {"indicador": "Personas procesadas", "valor": len(personas)},
        {"indicador": "Activos procesados", "valor": len(activos)},
        {"indicador": "Artículos procesados", "valor": len(articulos)},
        {"indicador": "Validaciones de datos", "valor": len(validacion_datos)},
        {"indicador": "Validaciones críticas", "valor": count_value(validacion_datos, "severidad", "Crítica")},
        {"indicador": "Alertas vigentes", "valor": len(alertas_generadas)},
        {"indicador": "Acciones pendientes vigentes", "valor": len(acciones_pendientes)},
        {"indicador": "Alertas críticas", "valor": count_value(alertas_generadas, "criticidad", "Crítica")},
        {"indicador": "Alertas altas", "valor": count_value(alertas_generadas, "criticidad", "Alta")},
        {"indicador": "Alertas medias", "valor": count_value(alertas_generadas, "criticidad", "Media")},
        {"indicador": "Alertas vencidas", "valor": count_value(acciones_pendientes, "alerta_vencida", "Sí")},
        {"indicador": "Nuevas detecciones", "valor": count_value(acciones_pendientes, "situacion_alerta", "Nueva detección")},
        {"indicador": "Alertas persistentes", "valor": count_value(acciones_pendientes, "situacion_alerta", "Persistente")},
        {"indicador": "Alertas no vigentes en historial", "valor": count_value(historial_alertas, "vigente_en_ultima_carga", "No")},
    ]
    dashboard_resumen = pd.DataFrame(dashboard_data)

    if not alertas_generadas.empty:
        alertas_por_regla = (
            alertas_generadas
            .groupby(["orden_prioridad", "codigo_regla", "tipo_alerta", "criticidad"], dropna=False)
            .size()
            .reset_index(name="cantidad_alertas")
            .sort_values("orden_prioridad")
        )
    else:
        alertas_por_regla = pd.DataFrame(columns=["orden_prioridad", "codigo_regla", "tipo_alerta", "criticidad", "cantidad_alertas"])

    if not acciones_pendientes.empty:
        alertas_por_estado_gestion = (
            acciones_pendientes
            .groupby(["estado_gestion"], dropna=False)
            .size()
            .reset_index(name="cantidad")
            .sort_values("cantidad", ascending=False)
        )
    else:
        alertas_por_estado_gestion = pd.DataFrame(columns=["estado_gestion", "cantidad"])

    trazabilidad_ejecucion = pd.DataFrame([{
        "timestamp_ejecucion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "archivo_fuente": archivo_fuente_nombre,
        "personas_procesadas": len(personas),
        "activos_procesados": len(activos),
        "articulos_procesados": len(articulos),
        "alertas_generadas": len(alertas_generadas),
        "acciones_pendientes": len(acciones_pendientes),
        "historial_alertas": len(historial_alertas),
        "umbral_costo_reparacion": umbral
    }])

    return dashboard_resumen, alertas_por_regla, alertas_por_estado_gestion, trazabilidad_ejecucion


def build_excel_bytes(
    dashboard_resumen,
    acciones_pendientes,
    validacion_datos,
    historial_alertas,
    reglas_aplicadas,
    trazabilidad_ejecucion,
    alertas_por_regla,
    alertas_por_estado_gestion
):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dashboard_resumen.to_excel(writer, index=False, sheet_name="resumen_ejecutivo", startrow=1)
        alertas_por_regla.to_excel(writer, index=False, sheet_name="resumen_ejecutivo", startrow=len(dashboard_resumen) + 5)
        alertas_por_estado_gestion.to_excel(writer, index=False, sheet_name="resumen_ejecutivo", startrow=len(dashboard_resumen) + len(alertas_por_regla) + 9)

        acciones_pendientes.to_excel(writer, index=False, sheet_name="acciones_pendientes")
        validacion_datos.to_excel(writer, index=False, sheet_name="validacion_datos")
        historial_alertas.to_excel(writer, index=False, sheet_name="historial_alertas")
        reglas_aplicadas.to_excel(writer, index=False, sheet_name="reglas_aplicadas")
        trazabilidad_ejecucion.to_excel(writer, index=False, sheet_name="trazabilidad_ejecucion")

    output.seek(0)

    wb = load_workbook(output)
    ws = wb["resumen_ejecutivo"]

    ws["A1"] = "Resumen ejecutivo"
    ws["A1"].font = Font(bold=True, size=14)

    row_alertas_regla = len(dashboard_resumen) + 5
    ws[f"A{row_alertas_regla}"] = "Alertas por regla"
    ws[f"A{row_alertas_regla}"].font = Font(bold=True, size=12)

    row_estado = len(dashboard_resumen) + len(alertas_por_regla) + 9
    ws[f"A{row_estado}"] = "Alertas por estado de gestión"
    ws[f"A{row_estado}"].font = Font(bold=True, size=12)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    for sheet in wb.worksheets:
        if sheet.title != "resumen_ejecutivo":
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for header_row in [2, row_alertas_regla + 1, row_estado + 1]:
        for cell in ws[header_row]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def process_data_control_hub(source_file, previous_file=None, umbral_costo_reparacion=0.70, source_filename="archivo_fuente.xlsx"):
    personas_raw, activos_raw, articulos_raw, reglas_raw = load_source_excel(source_file)
    personas, activos, articulos, reglas = standardize_and_validate_schema(personas_raw, activos_raw, articulos_raw, reglas_raw)
    personas, activos, articulos, reglas = normalize_data(personas, activos, articulos, reglas)

    validacion_datos = validate_data_quality(personas, activos, articulos)
    reglas_aplicadas = prepare_rules(reglas)
    base = cross_sources(personas, activos, articulos)
    alertas_generadas = apply_rules(base, reglas_aplicadas, umbral_costo_reparacion=umbral_costo_reparacion)

    gestion_anterior, historial_anterior = read_previous_file(previous_file)
    acciones_pendientes = generate_actions(alertas_generadas, gestion_anterior)
    historial_alertas = generate_history(acciones_pendientes, historial_anterior)

    dashboard_resumen, alertas_por_regla, alertas_por_estado_gestion, trazabilidad_ejecucion = generate_summaries(
        personas=personas,
        activos=activos,
        articulos=articulos,
        validacion_datos=validacion_datos,
        alertas_generadas=alertas_generadas,
        acciones_pendientes=acciones_pendientes,
        historial_alertas=historial_alertas,
        archivo_fuente_nombre=source_filename,
        umbral=umbral_costo_reparacion
    )

    return {
        "personas": personas,
        "activos": activos,
        "articulos": articulos,
        "validacion_datos": validacion_datos,
        "reglas_aplicadas": reglas_aplicadas,
        "alertas_generadas": alertas_generadas,
        "acciones_pendientes": acciones_pendientes,
        "historial_alertas": historial_alertas,
        "dashboard_resumen": dashboard_resumen,
        "alertas_por_regla": alertas_por_regla,
        "alertas_por_estado_gestion": alertas_por_estado_gestion,
        "trazabilidad_ejecucion": trazabilidad_ejecucion,
    }


def build_result_excel_from_result(result, acciones_pendientes_editadas=None):
    acciones = acciones_pendientes_editadas if acciones_pendientes_editadas is not None else result["acciones_pendientes"]

    # Regenerar resumen de estados por si se editó la gestión
    if not acciones.empty:
        alertas_por_estado_gestion = (
            acciones
            .groupby(["estado_gestion"], dropna=False)
            .size()
            .reset_index(name="cantidad")
            .sort_values("cantidad", ascending=False)
        )
    else:
        alertas_por_estado_gestion = result["alertas_por_estado_gestion"]

    return build_excel_bytes(
        dashboard_resumen=result["dashboard_resumen"],
        acciones_pendientes=acciones,
        validacion_datos=result["validacion_datos"],
        historial_alertas=result["historial_alertas"],
        reglas_aplicadas=result["reglas_aplicadas"],
        trazabilidad_ejecucion=result["trazabilidad_ejecucion"],
        alertas_por_regla=result["alertas_por_regla"],
        alertas_por_estado_gestion=alertas_por_estado_gestion
    )


# ============================================================
# ENVÍO POR CORREO
# ============================================================

def send_excel_by_email(
    excel_bytes,
    filename,
    correo_remitente,
    password_app,
    correo_destinatario,
    asunto,
    cuerpo,
    smtp_server="smtp.gmail.com",
    smtp_port=587
):
    mensaje = EmailMessage()
    mensaje["From"] = correo_remitente
    mensaje["To"] = correo_destinatario
    mensaje["Subject"] = asunto
    mensaje.set_content(cuerpo)

    mime_type, _ = mimetypes.guess_type(filename)
    if mime_type is None:
        mime_type = "application/octet-stream"

    maintype, subtype = mime_type.split("/", 1)

    mensaje.add_attachment(
        excel_bytes.getvalue(),
        maintype=maintype,
        subtype=subtype,
        filename=filename
    )

    with smtplib.SMTP(smtp_server, smtp_port) as smtp:
        smtp.starttls()
        smtp.login(correo_remitente, password_app)
        smtp.send_message(mensaje)
